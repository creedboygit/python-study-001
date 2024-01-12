import openpyxl

wb = openpyxl.Workbook()

ws = wb.create_sheet("Valletta")

ws["A1"] = "참가번호"
ws["B1"] = "성명"

ws["A2"] = 1
ws["B2"] = "발레타"

wb.save(
    r"D:\940 python_study\200 study\210 basic\pythonProject\study\real_crawler\files\참가자_data.xlsx"
)
