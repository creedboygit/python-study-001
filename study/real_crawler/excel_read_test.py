import openpyxl

fpath = r'D:\940 python_study\200 study\210 basic\pythonProject\study\real_crawler\files\참가자_data.xlsx'

wb = openpyxl.load_workbook(fpath)

ws = wb['Valletta']

ws['A3'] = 456
ws['B3'] = '크리드'

wb.save(fpath)

